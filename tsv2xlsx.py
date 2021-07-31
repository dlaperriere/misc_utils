#!/usr/bin/env python
"""
Create an excel file from a tsv file or a list of tsv files

Usage

     python tsv2xlsx.py -f tsv_file -x excel_filename.xlsx

     python tsv2xlsx.py -l tsv_list -x excel_filename.xlsx
     ls *.txt | python tsvlist2xlsx.py -l - -x excel_filename.xlsx

Copyright

David Laperriere dlaperriere@outlook.com
"""

import argparse
import csv
import os
import re
import textwrap

from openpyxl import Workbook

__version_info__ = (1, 0)
__version__ = '.'.join(map(str, __version_info__))
__author__ = "David Laperriere dlaperriere@outlook.com"


def build_argparser():
    """ Build command line arguments parser """
    parser = argparse.ArgumentParser(
        prog=__file__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
        description=textwrap.dedent('''
      Create an excel file from a list of tsv files
                               '''),
        epilog=textwrap.dedent('''
    Examples

       python tsv2xlsx.py -f tsv_file -x excel_filename.xlsx

       python tsv2xlsx.py -l tsv_list -x excel_filename.xlsx
       ls *.txt | python tsvlist2xlsx.py -l - -x excel_filename.xlsx

        ''')
    )
    parser.add_argument('-f', '--file',
                        type=str,
                        required=False,
                        default=None,
                        help='tsv file')
    parser.add_argument('-l', '--list',
                        type=argparse.FileType('r'),
                        required=False,
                        default=None,
                        help='list of tsv files (one per line)')
    parser.add_argument('-x', '--excel',
                        type=str,
                        required=True,
                        help='excel file name')

    parser.add_argument('-v', '--version', action='version',
                        version="%(prog)s v" + __version__)

    return parser


def excel_add_tsv(filename, title, workbook):
    """
    Add content of a tsv file to an excel worksheet

        parameters
         - filename: file to add
         - title: worksheet name
         - workbook: openpyxl workbook
    """
    # http://openpyxl.readthedocs.io/en/2.3.3/_modules/openpyxl/workbook/child.html
    invalid_char = r'[\\*?:/\[\]]'
    title = re.sub(invalid_char, '_', title)
    max_length = 31
    if len(title) > max_length:
        title = title[0:max_length - 1]

    ws = workbook.create_sheet(title=title)

    with open(filename, "rU") as tab_file:
        tab_reader = csv.reader(tab_file, delimiter='\t')
        for idx, line in enumerate(tab_reader):
            for column in range(len(line)):
                _ = ws.cell(row=idx + 1, column=column + 1, value=line[column])


def excel_from_tsv(tsv_files, excel_name):
    """
    Create an excel from a list of tsv file

        parameters
         - tsv_files: list of tsv files
         - excel_name: excel file name
    """
    wb = Workbook()
    ws = wb.active

    ws.title = "Files"
    row = 1
    ws.cell(column=1, row=row, value="File")
    for tsv_file in tsv_files:
        row += 1
        ws.cell(column=1, row=row, value=tsv_file)

    for tsv_file in tsv_files:
        filename = os.path.basename(tsv_file)
        print("  - add {} ".format(filename))
        excel_add_tsv(tsv_file, filename, wb)

    wb.save(excel_name)


def main():
    """ Main: parse arguments and create excel file """

    # parse command line arguments
    parser = build_argparser()
    pyargs = parser.parse_args()

    if pyargs.file is None and pyargs.list is None:
        parser.exit(
            message="must provide tsv file (-f) or tsv list (-l)...", status=-1)

    tsv_list = list()
    if pyargs.file is not None:
        tsv_list.append(pyargs.file)
    else:
        for tsv in pyargs.list:
            tsv_list.append(tsv.rstrip())

    # skip dir and invalid path
    tsv_files = list()
    for tsv_name in tsv_list:
        if os.path.isfile(tsv_name):
            tsv_files.append(tsv_name)
        else:
            print("  - skip \'{}\'".format(tsv_name))

    # create excel file
    excel_name = pyargs.excel
    if not excel_name.endswith('.xlsx'):
        excel_name += '.xlsx'

    if len(tsv_files) > 0:
        excel_from_tsv(tsv_files, excel_name)


if __name__ == "__main__":
    main()
    print("\nDone")
    exit(0)
